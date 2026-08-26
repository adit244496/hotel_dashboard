"""Parser for the IHCL/Taj "MIS Financial Book" workbook family.

Covers both variants seen in the wild:

* named sheets — ``1.1MainP&LSummary``, ``1.2StatementofRoomRevenue(`` …
  (also with spaces: ``1.1 Main P&L Summary_ICP_M``)
* generic sheets — ``Table 1``, ``Table 3`` … produced by some exports

Both carry identical line-item labels, so a single content-driven parser
handles them. Nothing is addressed by fixed row or column numbers.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from app.parsers.base import OutletRow, ParseResult, ParserError, SegmentRow
from app.parsers.mapping import (
    ACCUMULATING,
    FNB_COSTS,
    FNB_SALES,
    OUTLET_TOTAL_LABELS,
    SEGMENT_ROW_METRICS,
    SEGMENTS,
    match_metric,
    match_outlet_group,
)
from app.parsers.utils import (
    ColumnGroup,
    find_header_row,
    iter_data_rows,
    month_from_label,
    norm,
    read_column_groups,
    to_float,
)

PARSER_NAME = "ihcl_financial_book"

_ENTITY = re.compile(r"\b(E_\d{3,6})\b", re.IGNORECASE)
_COVER_KINDS = {"covers_resident", "covers_non_resident", "covers_total"}


# --------------------------------------------------------------------------- #
# Sheet discovery
# --------------------------------------------------------------------------- #
def _column_texts(ws, col: int, max_row: int = 80) -> list[str]:
    limit = min(ws.max_row, max_row)
    return [norm(ws.cell(r, col).value) for r in range(1, limit + 1)]


def _label_column(ws, header_row: int, default: int = 1) -> int:
    """Pick the column holding line-item labels (usually A, occasionally B)."""
    best, best_count = default, 0
    for col in (1, 2):
        count = sum(
            1
            for r in range(header_row + 1, min(ws.max_row, header_row + 60) + 1)
            if isinstance(ws.cell(r, col).value, str)
            and ws.cell(r, col).value.strip()
        )
        if count > best_count:
            best, best_count = col, count
    return best


def _name_bonus(sheet_name: str, hints: tuple[str, ...]) -> int:
    key = norm(sheet_name).replace(" ", "")
    return 100 if any(h.replace(" ", "") in key for h in hints) else 0


def _period_score(ws) -> tuple[int, int]:
    """(number of usable period groups, number of scenario columns)."""
    header_row = find_header_row(ws)
    if header_row is None:
        return 0, 0
    groups = read_column_groups(ws, header_row)
    periods = _period_groups(groups)
    return len(periods), sum(len(g.columns) for _pt, _m, g in periods)


def _best_sheet(wb, predicate, hints: tuple[str, ...]):
    """Pick the richest sheet matching ``predicate``.

    Workbooks often contain several sheets that satisfy a naive text test — a
    'PL Variance MTD' extract sits alongside the real '1.1 Main P&L Summary',
    for instance — so candidates are scored on name and on how many period
    columns they actually expose.
    """
    best, best_score = None, None
    for name in wb.sheetnames:
        ws = wb[name]
        if not predicate(ws):
            continue
        groups, columns = _period_score(ws)
        score = (_name_bonus(name, hints), groups, columns)
        if best_score is None or score > best_score:
            best, best_score = ws, score
    return best


def _is_pl_sheet(ws) -> bool:
    texts = set(_column_texts(ws, 1)) | set(_column_texts(ws, 2))
    return (
        "turnover" in texts
        and ("room income" in texts or "ebitda" in texts)
        and find_header_row(ws) is not None
    )


def _is_segment_sheet(ws) -> bool:
    texts = set(_column_texts(ws, 1))
    return "corporate" in texts and any(
        t.startswith("rooms occupied per day") for t in texts
    )


def _is_fnb_category_sheet(ws) -> bool:
    # Joined rather than set-matched: some exports pack the whole sales block
    # into one newline-separated cell, which _expand_packed_row unpacks later.
    joined = " ".join(_column_texts(ws, 1))
    return "food sales" in joined and "total sales" in joined


def _find_pl_sheet(wb):
    return _best_sheet(wb, _is_pl_sheet, ("1.1", "main p&l", "main p & l"))


def _find_segment_sheet(wb):
    return _best_sheet(wb, _is_segment_sheet, ("1.2", "room revenue"))


def _find_fnb_category_sheet(wb):
    return _best_sheet(wb, _is_fnb_category_sheet, ("1.5", "f&b category"))


def _find_outlet_sheet(wb):
    """The outlet sheet is the one whose header groups include cover splits.

    Some exports truncate the sheet after the first couple of cover groups, so
    only the revenue group plus *any* cover group is required.
    """
    best, best_score = (None, None, None), None
    for name in wb.sheetnames:
        ws = wb[name]
        header_row = find_header_row(ws, max_row=12)
        if header_row is None:
            continue
        groups = read_column_groups(ws, header_row)
        kinds = {match_outlet_group(g.label) for g in groups}
        kinds.discard(None)
        if "revenue" not in kinds or not (kinds & _COVER_KINDS):
            continue
        # Reject the forecast twin (1.4), which has no current-year actual.
        revenue_group = next(
            g for g in groups if match_outlet_group(g.label) == "revenue"
        )
        if "ACT" not in revenue_group.columns:
            continue
        score = (_name_bonus(name, ("1.3", "outlet")), len(kinds))
        if best_score is None or score > best_score:
            best, best_score = (ws, header_row, groups), score
    return best


# --------------------------------------------------------------------------- #
# Header helpers
# --------------------------------------------------------------------------- #
def _period_of(group: ColumnGroup) -> tuple[str, int] | None:
    """Classify a header group as an MTD or YTD period and pull out the month."""
    key = norm(group.label)
    month = month_from_label(key)
    if month is None:
        return None
    if "ytd" in key or "year to date" in key:
        return "YTD", month
    # Ranges such as 'April - December' are cumulative, not a single month.
    if "-" in key or " to " in key:
        return "YTD", month
    return "MTD", month


def _period_groups(groups: list[ColumnGroup]) -> list[tuple[str, int, ColumnGroup]]:
    out = []
    for g in groups:
        period = _period_of(g)
        if period and g.columns:
            out.append((period[0], period[1], g))
    return out


def _entity_code(ws) -> str | None:
    for r in range(1, min(ws.max_row, 12) + 1):
        for c in range(1, min(ws.max_column, 4) + 1):
            m = _ENTITY.search(str(ws.cell(r, c).value or ""))
            if m:
                return m.group(1).upper()
    return None


def _hotel_hint(wb) -> str | None:
    """Free-text hotel name, printed at the top of some sheets."""
    for name in wb.sheetnames:
        ws = wb[name]
        for r in range(1, min(ws.max_row, 8) + 1):
            for c in range(1, min(ws.max_column, 3) + 1):
                value = ws.cell(r, c).value
                if isinstance(value, str):
                    text = value.strip()
                    low = text.lower()
                    if len(text) > 8 and (
                        "taj" in low
                        or "hotel" in low
                        or "resort" in low
                        or "kutir" in low
                    ):
                        return text
    return None


# --------------------------------------------------------------------------- #
# Section parsers
# --------------------------------------------------------------------------- #
def _parse_pl(ws, result: ParseResult) -> None:
    header_row = find_header_row(ws)
    if header_row is None:
        raise ParserError("Could not locate the 'Particulars' header row.")
    groups = read_column_groups(ws, header_row)
    periods = _period_groups(groups)
    if not periods:
        raise ParserError("No 'Dec' / 'YTD Dec' style period columns were found.")

    label_col = _label_column(ws, header_row)
    value_cols = sorted({c for _, _, g in periods for c in g.columns.values()})
    months = {month for _, month, _ in periods}
    result.month = max(months) if len(months) == 1 else _dominant_month(periods)

    seen_metrics: set[str] = set()
    for row in iter_data_rows(ws, header_row + 1, value_cols, label_col=label_col):
        if not row.has_numbers:
            continue
        metric = None
        for candidate in row.candidate_labels():
            metric = match_metric(candidate)
            if metric:
                break
        if not metric:
            continue
        accumulate = metric in ACCUMULATING
        seen_metrics.add(metric)
        for period_type, _month, group in periods:
            for scenario, col in group.columns.items():
                value = row.values.get(col)
                if value is None:
                    continue
                result.add(period_type, scenario, metric, value, accumulate)

    if "turnover" not in seen_metrics and "room_income" not in seen_metrics:
        raise ParserError("Sheet does not look like a P&L summary (no Turnover row).")

    inv = result.bucket("MTD", "ACT").get("room_inventory")
    if inv:
        result.room_inventory = inv
    result.sheets_used.append(ws.title)


def _dominant_month(periods) -> int:
    counts: dict[int, int] = {}
    for _pt, month, _g in periods:
        counts[month] = counts.get(month, 0) + 1
    return max(counts, key=lambda m: (counts[m], m))


FNB_SECTIONS = ("sales", "raw material cost %", "raw material cost")


def _greedy_split(text: str, vocabulary) -> list[str] | None:
    """Split a run-together label string using a known label vocabulary.

    Wrapped cells lose some newlines, so 'Soft Drinks & Mineral Water Sales'
    and 'Beer Sales' can end up on one line. Matching longest-known-label-first
    recovers the original boundaries, or returns None if anything is left over.
    """
    key = norm(text)
    known = sorted(vocabulary, key=len, reverse=True)
    out: list[str] = []
    pos = 0
    while pos < len(key):
        if key[pos] == " ":
            pos += 1
            continue
        for candidate in known:
            if key.startswith(candidate, pos):
                out.append(candidate)
                pos += len(candidate)
                break
        else:
            return None
    return out or None


def _expand_packed_row(ws, row: int, label_col: int, value_cols: list[int]):
    """Recover rows that an export collapsed into newline-packed cells.

    Some workbooks emit a whole block as a single row whose label cell holds
    ``"Sales\\nFood Sales\\nBeer Sales..."`` and whose value cells hold the
    matching numbers, one per line. Unpacking restores ordinary rows; when the
    pieces cannot be lined up confidently the row is left alone.
    """
    raw_label = ws.cell(row, label_col).value
    if not isinstance(raw_label, str) or "\n" not in raw_label:
        return None

    label_lines = [line.strip() for line in raw_label.split("\n") if line.strip()]
    if len(label_lines) < 2:
        return None

    columns: dict[int, list[float | None]] = {}
    for col in value_cols:
        raw = ws.cell(row, col).value
        if not isinstance(raw, str) or "\n" not in raw:
            continue
        columns[col] = [to_float(part) for part in raw.split("\n")]
    if not columns:
        return None

    counts = {len(v) for v in columns.values()}
    if len(counts) != 1:
        return None
    count = counts.pop()

    section = ""
    remaining = label_lines
    for known_section in FNB_SECTIONS:
        if norm(label_lines[0]) == known_section:
            section, remaining = label_lines[0], label_lines[1:]
            break

    labels: list[str] | None = None
    if section:
        vocabulary = (
            FNB_COSTS if norm(section).startswith("raw material cost") else FNB_SALES
        )
        if "%" not in section:
            labels = _greedy_split(" ".join(remaining), vocabulary)
    if labels is None or len(labels) != count:
        labels = remaining if len(remaining) == count else None
    if labels is None:
        labels = label_lines if len(label_lines) == count else None
    if labels is None:
        return None

    out = []
    for idx, label in enumerate(labels):
        values = {
            col: parts[idx]
            for col, parts in columns.items()
            if parts[idx] is not None
        }
        out.append((section, label, values))
    return out


def _parse_fnb_categories(ws, result: ParseResult) -> None:
    header_row = find_header_row(ws)
    if header_row is None:
        return
    groups = read_column_groups(ws, header_row)
    periods = _period_groups(groups)
    if not periods:
        return
    label_col = _label_column(ws, header_row)
    value_cols = sorted({c for _, _, g in periods for c in g.columns.values()})

    # Flatten the sheet into (section, label, values) triples first so packed
    # and ordinary rows can be handled the same way.
    entries: list[tuple[str, str, dict[int, float]]] = []
    section = ""
    for row in iter_data_rows(ws, header_row + 1, value_cols, label_col=label_col):
        packed = _expand_packed_row(ws, row.row, label_col, value_cols)
        if packed:
            for packed_section, label, values in packed:
                entries.append((packed_section or section, label, values))
            if packed[0][0]:
                section = packed[0][0]
            continue
        if not row.has_numbers:
            section = row.label
            continue
        entries.append((section, row.label, row.values))

    for entry_section, label, values in entries:
        key = norm(label)
        section_key = norm(entry_section)
        if "%" in section_key:
            continue  # the ratio block restates what we already store
        if section_key.startswith("raw material cost"):
            metric = FNB_COSTS.get(key)
        else:
            metric = FNB_SALES.get(key)
        if not metric:
            continue
        for period_type, _month, group in periods:
            for scenario, col in group.columns.items():
                value = values.get(col)
                if value is not None:
                    result.bucket(period_type, scenario).setdefault(metric, value)
    result.sheets_used.append(ws.title)


def _parse_segments(ws, result: ParseResult) -> None:
    header_row = find_header_row(ws)
    if header_row is None:
        return
    groups = read_column_groups(ws, header_row)
    periods = _period_groups(groups)
    if not periods:
        return
    label_col = _label_column(ws, header_row)
    value_cols = sorted({c for _, _, g in periods for c in g.columns.values()})

    # (period, scenario, segment) -> partially filled row
    rows: dict[tuple[str, str, str], SegmentRow] = {}
    current: str | None = None

    for row in iter_data_rows(ws, header_row + 1, value_cols, label_col=label_col):
        key = norm(row.label)
        if not row.has_numbers:
            current = SEGMENTS.get(key)
            continue
        if current is None:
            continue
        field = SEGMENT_ROW_METRICS.get(key)
        if field is None:
            for prefix, name in SEGMENT_ROW_METRICS.items():
                if key.startswith(prefix.split(" (")[0]):
                    field = name
                    break
        if field is None:
            continue
        for period_type, _month, group in periods:
            for scenario, col in group.columns.items():
                value = row.values.get(col)
                if value is None:
                    continue
                bucket = rows.setdefault(
                    (period_type, scenario, current),
                    SegmentRow(
                        period_type=period_type, scenario=scenario, segment=current
                    ),
                )
                if getattr(bucket, field) is None:
                    setattr(bucket, field, value)

    result.segments.extend(rows.values())
    result.sheets_used.append(ws.title)


def _parse_outlets(ws, header_row: int, groups: list[ColumnGroup], result: ParseResult) -> None:
    label_col = _label_column(ws, header_row)
    metric_groups = [
        (kind, g) for g in groups if (kind := match_outlet_group(g.label)) is not None
    ]
    value_cols = sorted({c for _m, g in metric_groups for c in g.columns.values()})

    rows: dict[tuple[str, str], OutletRow] = {}
    totals: dict[str, float] = {}

    for row in iter_data_rows(ws, header_row + 1, value_cols, label_col=label_col):
        if not row.has_numbers:
            continue
        key = norm(row.label)
        is_total = key in OUTLET_TOTAL_LABELS
        for metric, group in metric_groups:
            for scenario, col in group.columns.items():
                value = row.values.get(col)
                if value is None:
                    continue
                if is_total:
                    if metric == "revenue" and scenario == "ACT":
                        totals["revenue"] = value
                    continue
                bucket = rows.setdefault(
                    (scenario, row.label.strip()),
                    OutletRow(
                        period_type="MTD", scenario=scenario, outlet=row.label.strip()
                    ),
                )
                if getattr(bucket, metric) is None:
                    setattr(bucket, metric, value)

    # The outlet statement has no period header of its own. Decide MTD vs YTD by
    # checking which F&B income figure its grand total agrees with.
    period_type = "MTD"
    total = totals.get("revenue")
    if total:
        mtd = result.bucket("MTD", "ACT").get("fnb_income")
        ytd = result.bucket("YTD", "ACT").get("fnb_income")
        candidates = [(abs(total - v), p) for p, v in (("MTD", mtd), ("YTD", ytd)) if v]
        if candidates:
            diff, period_type = min(candidates)
            reference = mtd if period_type == "MTD" else ytd
            if diff > max(1.0, 0.02 * total):
                result.warnings.append(
                    f"Outlet statement totals {total:,.1f}L but the P&L shows "
                    f"{reference:,.1f}L of {period_type} F&B income — a gap of "
                    f"{diff:,.1f}L in the source workbook. Outlet rows stored as "
                    f"{period_type}; outlet revenue will not sum to F&B revenue."
                )
    for row in rows.values():
        row.period_type = period_type
        if row.covers_total is None:
            parts = [row.covers_resident, row.covers_non_resident]
            if any(p is not None for p in parts):
                row.covers_total = sum(p for p in parts if p is not None)
        # Truncated exports stop before the APC group; revenue (in lakhs) over
        # covers gives the same figure.
        if row.apc is None and row.revenue and row.covers_total:
            row.apc = row.revenue * 100_000 / row.covers_total

    result.outlets.extend(rows.values())
    result.sheets_used.append(ws.title)


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #
def detect(path: Path) -> bool:
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception:
        return False
    try:
        return _find_pl_sheet(wb) is not None
    finally:
        wb.close()


def parse(path: Path) -> ParseResult:
    wb = load_workbook(path, data_only=True)
    try:
        result = ParseResult(parser=PARSER_NAME)
        pl_ws = _find_pl_sheet(wb)
        if pl_ws is None:
            raise ParserError(
                "No Main P&L Summary sheet found (expected rows 'Turnover' and "
                "'Room Income')."
            )
        _parse_pl(pl_ws, result)
        result.entity_code = _entity_code(pl_ws)
        result.hotel_hint = _hotel_hint(wb)

        header_row = find_header_row(pl_ws)
        from app.parsers.utils import detect_fiscal_year

        result.fiscal_year = detect_fiscal_year(pl_ws, header_row)

        fnb_ws = _find_fnb_category_sheet(wb)
        if fnb_ws is not None:
            _parse_fnb_categories(fnb_ws, result)
        else:
            result.warnings.append("F&B category sheet not found; cost split omitted.")

        seg_ws = _find_segment_sheet(wb)
        if seg_ws is not None:
            _parse_segments(seg_ws, result)
        else:
            result.warnings.append("Room revenue segment sheet not found.")

        out_ws, out_header, out_groups = _find_outlet_sheet(wb)
        if out_ws is not None:
            _parse_outlets(out_ws, out_header, out_groups, result)
        else:
            result.warnings.append("Outlet-wise F&B sheet not found.")

        return result
    finally:
        wb.close()


__all__ = ["PARSER_NAME", "detect", "parse", "to_float"]
