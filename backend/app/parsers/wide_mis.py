"""Parser for the bespoke wide "MIS" workbook (one sheet, a block per month).

Layout: a header row carries month captions ('DECEMBER', 'APRIL - DECEMBER')
every N columns; the row beneath labels each column within the block
('FY26 Act.', 'FY26 Bud.', 'FY25 Act.', 'VarBud', ...). Line items run down a
label column in stacked sections — P&L, performance parameters, F&B workings,
room revenue by segment, and outlet-wise analysis.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from app.parsers.base import OutletRow, ParseResult, ParserError, SegmentRow
from app.parsers.mapping import (
    ACCUMULATING,
    OUTLET_ROW_METRICS,
    OUTLET_TOTAL_LABELS,
    SEGMENT_ROW_METRICS_SQUASHED,
    WORKINGS_METRICS,
    match_metric,
    match_segment,
    squash,
)
from app.parsers.utils import MONTHS, norm, to_float

PARSER_NAME = "wide_mis"

_FY = re.compile(r"fy\s*(\d{2})", re.IGNORECASE)

# Section captions that switch the parser into a different reading mode.
_SECTION_PL = ("profit and loss", "profit & loss")
_SECTION_PARAMS = ("performance parameters",)
_SECTION_WORKINGS = ("workings",)
_SECTION_SEGMENTS = ("room revenue",)
_SECTION_OUTLETS = ("outletwise analysis", "outlet wise analysis")


def _is_month_word(text: str) -> bool:
    return norm(text) in MONTHS


def _find_month_header(ws, max_row: int = 12) -> tuple[int, list[tuple[str, int]]] | None:
    """Locate the row of month captions and the column each block starts at."""
    for r in range(1, min(ws.max_row, max_row) + 1):
        blocks: list[tuple[str, int]] = []
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            if not isinstance(value, str) or not value.strip():
                continue
            words = [w for w in re.split(r"[^A-Za-z]+", value) if w]
            if words and all(_is_month_word(w) for w in words):
                blocks.append((value.strip(), c))
        if len(blocks) >= 6:
            return r, blocks
    return None


def _label_column(ws, start_row: int, blocks_start: int) -> int:
    best, best_count = 1, 0
    for col in range(1, min(blocks_start, 4)):
        count = sum(
            1
            for r in range(start_row, min(ws.max_row, start_row + 120) + 1)
            if isinstance(ws.cell(r, col).value, str) and ws.cell(r, col).value.strip()
        )
        if count > best_count:
            best, best_count = col, count
    return best


def _classify_block(ws, header_row: int, start: int, end: int) -> dict[str, int]:
    """Map ACT / BUD / LY onto the columns inside one month block."""
    entries: list[tuple[int, str, int]] = []  # (col, kind, fy)
    for c in range(start, end + 1):
        parts = []
        for r in range(header_row, min(header_row + 3, ws.max_row) + 1):
            value = ws.cell(r, c).value
            if isinstance(value, str) and value.strip():
                parts.append(norm(value))
        sig = " ".join(parts)
        if not sig or "var" in sig:
            continue
        fy_match = _FY.search(sig)
        if not fy_match:
            continue
        fy = int(fy_match.group(1))
        if "bud" in sig:
            entries.append((c, "BUD", fy))
        elif "act" in sig:
            entries.append((c, "ACT", fy))
    if not entries:
        return {}
    current = max(fy for _c, _k, fy in entries)
    columns: dict[str, int] = {}
    for col, kind, fy in entries:
        if kind == "BUD":
            scenario = "BUD"
        elif fy == current:
            scenario = "ACT"
        elif fy == current - 1:
            scenario = "LY"
        else:
            continue  # older comparatives (FY24, FY23) are not modelled
        columns.setdefault(scenario, col)
    return columns


def _period_of(caption: str) -> tuple[str, int] | None:
    words = [w for w in re.split(r"[^A-Za-z]+", caption) if w]
    months = [MONTHS[norm(w)] for w in words if norm(w) in MONTHS]
    if not months:
        return None
    if len(months) > 1 or "-" in caption:
        return "YTD", months[-1]
    return "MTD", months[0]


def _fiscal_year(ws, header_row: int, blocks: list[tuple[str, int]]) -> str | None:
    years = set()
    for _caption, start in blocks:
        for r in range(header_row, min(header_row + 3, ws.max_row) + 1):
            for c in range(start, min(start + 8, ws.max_column) + 1):
                m = _FY.search(str(ws.cell(r, c).value or ""))
                if m:
                    years.add(int(m.group(1)))
    if not years:
        return None
    current = max(years)
    return f"20{current - 1}-{current}"


def _find_mis_sheet(wb):
    best, best_score = None, None
    for name in wb.sheetnames:
        ws = wb[name]
        header = _find_month_header(ws)
        if header is None:
            continue
        header_row, blocks = header
        label_col = _label_column(ws, header_row + 1, blocks[0][1])
        labels = {
            norm(ws.cell(r, label_col).value)
            for r in range(header_row, min(ws.max_row, header_row + 120) + 1)
        }
        if "turnover" not in labels:
            continue
        score = (len(blocks), ws.max_row)
        if best_score is None or score > best_score:
            best, best_score = (ws, header_row, blocks, label_col), score
    return best


def detect(path: Path) -> bool:
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return False
    try:
        return _find_mis_sheet(wb) is not None
    finally:
        wb.close()


def parse(path: Path, target_month: int | None = None) -> ParseResult:
    """Parse the block for ``target_month``.

    This workbook holds every month of the fiscal year — actuals for months
    already closed and projections for the rest — so the caller must say which
    month it is uploading. Without it we fall back to the latest month carrying
    actuals and flag the guess.
    """
    wb = load_workbook(path, data_only=True)
    try:
        found = _find_mis_sheet(wb)
        if found is None:
            raise ParserError("No wide MIS sheet with monthly blocks was found.")
        ws, header_row, blocks, label_col = found

        result = ParseResult(parser=PARSER_NAME)
        result.fiscal_year = _fiscal_year(ws, header_row, blocks)
        result.hotel_hint = _hotel_hint(ws, label_col)

        # Resolve each month block into (period_type, month, {scenario: col}).
        periods: list[tuple[str, int, dict[str, int]]] = []
        for idx, (caption, start) in enumerate(blocks):
            end = blocks[idx + 1][1] - 1 if idx + 1 < len(blocks) else ws.max_column
            period = _period_of(caption)
            if period is None:
                continue
            columns = _classify_block(ws, header_row, start, end)
            if columns:
                periods.append((period[0], period[1], columns))
        if not periods:
            raise ParserError("Monthly blocks carry no Act / Bud / LY columns.")

        available = sorted({month for _pt, month, _c in periods})
        if target_month is None:
            result.month = _reported_month(ws, header_row, periods, label_col)
            result.warnings.append(
                "This workbook contains every month of the year. Month "
                f"{result.month} was inferred from the latest actuals — check "
                "it matches the month you are uploading."
            )
        elif target_month not in available:
            raise ParserError(
                f"The workbook has no block for month {target_month}. "
                f"Months present: {', '.join(str(m) for m in available)}."
            )
        else:
            result.month = target_month

        active = [p for p in periods if p[1] == result.month]
        if not active:
            raise ParserError(f"No columns found for month {result.month}.")

        _read_sections(ws, header_row, label_col, active, result)
        result.sheets_used.append(ws.title)
        inventory = result.bucket("MTD", "ACT").get("room_inventory")
        if inventory:
            result.room_inventory = inventory
        return result
    finally:
        wb.close()


def _hotel_hint(ws, label_col: int) -> str | None:
    for r in range(1, 6):
        for c in (label_col, label_col + 1, 1, 2):
            value = ws.cell(r, c).value
            if isinstance(value, str) and len(value.strip()) > 8:
                low = value.lower()
                if "taj" in low or "hotel" in low or "resort" in low or "kutir" in low:
                    return value.strip()
    return None


def _reported_month(ws, header_row: int, periods, label_col: int) -> int:
    """The month the file reports on: the latest one carrying actuals.

    A projection workbook holds every month of the year, most of them empty, so
    the reported month is the last with a non-zero turnover actual.
    """
    turnover_row = None
    for r in range(header_row, min(ws.max_row, header_row + 120) + 1):
        if norm(ws.cell(r, label_col).value) == "turnover":
            turnover_row = r
            break
    if turnover_row is None:
        return max(month for _pt, month, _c in periods)

    best = None
    for period_type, month, columns in periods:
        if period_type != "MTD":
            continue
        col = columns.get("ACT")
        if col is None:
            continue
        value = to_float(ws.cell(turnover_row, col).value)
        if value:
            best = month if best is None else _later_fiscal(best, month)
    return best or max(month for _pt, month, _c in periods)


def _later_fiscal(a: int, b: int) -> int:
    order = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
    return a if order.index(a) > order.index(b) else b


def _read_sections(ws, header_row: int, label_col: int, periods, result: ParseResult) -> None:
    mode = "pl"
    current_segment: str | None = None
    current_outlet: str | None = None
    segments: dict[tuple[str, str, str], SegmentRow] = {}
    outlets: dict[tuple[str, str, str], OutletRow] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(r, label_col).value
        if not isinstance(raw, str) or not raw.strip():
            continue
        label = raw.strip()
        key = norm(label)

        if any(s in key for s in _SECTION_OUTLETS):
            mode, current_outlet = "outlets", None
            continue
        if any(s in key for s in _SECTION_SEGMENTS):
            mode, current_segment = "segments", None
            continue
        if any(s in key for s in _SECTION_WORKINGS):
            mode = "workings"
            continue
        if any(key.startswith(s) for s in _SECTION_PL) or any(
            s in key for s in _SECTION_PARAMS
        ):
            mode = "pl"
            continue

        values = {
            col: to_float(ws.cell(r, col).value)
            for _pt, _m, columns in periods
            for col in columns.values()
        }
        has_numbers = any(v is not None for v in values.values())

        if mode == "pl":
            metric = match_metric(label)
            if metric and has_numbers:
                accumulate = metric in ACCUMULATING
                for period_type, _m, columns in periods:
                    for scenario, col in columns.items():
                        value = values.get(col)
                        if value is not None:
                            result.add(period_type, scenario, metric, value, accumulate)

        elif mode == "workings":
            metric = WORKINGS_METRICS.get(squash(label))
            if metric and has_numbers:
                for period_type, _m, columns in periods:
                    for scenario, col in columns.items():
                        value = values.get(col)
                        if value is not None:
                            result.bucket(period_type, scenario).setdefault(metric, value)

        elif mode == "segments":
            if not has_numbers:
                current_segment = match_segment(label)
                continue
            field = SEGMENT_ROW_METRICS_SQUASHED.get(squash(label))
            if current_segment is None or field is None:
                continue
            for period_type, _m, columns in periods:
                for scenario, col in columns.items():
                    value = values.get(col)
                    if value is None:
                        continue
                    row = segments.setdefault(
                        (period_type, scenario, current_segment),
                        SegmentRow(
                            period_type=period_type,
                            scenario=scenario,
                            segment=current_segment,
                        ),
                    )
                    if getattr(row, field) is None:
                        setattr(row, field, value)

        elif mode == "outlets":
            field = OUTLET_ROW_METRICS.get(squash(label))
            if field is None:
                if not has_numbers:
                    current_outlet = label
                continue
            if current_outlet is None:
                continue
            for period_type, _m, columns in periods:
                for scenario, col in columns.items():
                    value = values.get(col)
                    if value is None:
                        continue
                    row = outlets.setdefault(
                        (period_type, scenario, current_outlet),
                        OutletRow(
                            period_type=period_type,
                            scenario=scenario,
                            outlet=current_outlet,
                        ),
                    )
                    if getattr(row, field) is None:
                        setattr(row, field, value)

    # 'TOTAL' captions are roll-up rows, not a segment or an outlet.
    result.segments.extend(
        row for (_pt, _sc, seg), row in segments.items() if seg is not None
    )
    result.outlets.extend(
        row
        for row in outlets.values()
        if norm(row.outlet) not in OUTLET_TOTAL_LABELS and row.revenue is not None
    )


__all__ = ["PARSER_NAME", "detect", "parse"]
